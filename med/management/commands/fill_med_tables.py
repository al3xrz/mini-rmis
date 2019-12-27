from django.core.management.base import BaseCommand, CommandError
from django.apps import apps
import datetime
from med.models import *
import openpyxl





class Command(BaseCommand):

    def fill_profile(self, filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb['КСГ']
        row = 2
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            excel_row = [str(sheet.cell(row=row, column=col).value) for col in range(5,7)]
            print("%s %s"%(excel_row[0], excel_row[1]))
            try:
                Profile.objects.get(code=excel_row[0])
            except:
                p = Profile(code=excel_row[0], name=excel_row[1])
                p.save()
            row += 1
        wb.close()


    def fill_ksg_items(self, filename, st_type):

        wb = openpyxl.load_workbook(filename)
        sheet = wb['КСГ']
        row = 2
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            excel_row = [str(sheet.cell(row=row, column=col).value) for col in range(2,7)]
            print("%s %s"%(row, excel_row))
            ksg = KSG(
                st_type = st_type,
                code = excel_row[0],
                name = excel_row[1],
                kz = excel_row[2],
                #profile_code = excel_row[3],
                profile=Profile.objects.get(code=excel_row[3])
            )
            ksg.save()
            row += 1
        wb.close()



    def fill_mkb_items(self, filename):

        wb = openpyxl.load_workbook(filename)
        sheet = wb['МКБ 10']
        row = 2
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            excel_row = [str(sheet.cell(row=row, column=col).value) for col in range(1,3)]
            ksg_row = [str(sheet.cell(row=row, column=col).value) for col in range(3,19) if str(sheet.cell(row=row, column=col).value)!='None']
            print(excel_row)
            mkb = MKB(code=excel_row[0], diagnosis=excel_row[1])
            mkb.save()
            for ksg_code in ksg_row:
                print(ksg_code)
                #print(KSG.objects.get(code=ksg_code))
                mkb.ksg.add(KSG.objects.get(code=ksg_code))
            row += 1
        wb.close()


    def fill_schemes(self, filename, st_type):
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Схемы лекарственной терапии']
        row = 2
        while True:
            mark_cell = sheet.cell(row=row, column=1).value
            if mark_cell is None:
                break
            excel_row = [str(sheet.cell(row=row, column=col).value) for col in range(1,8)]
            print("%s %s"%(row, excel_row))
            try:
                ksg = KSG.objects.get(code=excel_row[4])
            except:
                ksg = None

            scheme = Scheme(
                st_type = st_type,
                code = excel_row[0],
                mnn = excel_row[1],
                name = excel_row[2],
                col = excel_row[3],
                ksg = ksg,
                sign = excel_row[5],
                comments = excel_row[6]
            )
            print(scheme)
            scheme.save()
            row += 1
        wb.close()

















    def handle(self, *args, **options):

        # Profile.objects.all().delete()
        # self.fill_profile('s.xlsx')
        # self.fill_profile('d.xlsx')
        #
        # KSG.objects.all().delete()
        # self.fill_ksg_items('d.xlsx', "Дневной")
        # self.fill_ksg_items('s.xlsx', "Круглосуточный")
        #
        # MKB.objects.all().delete()
        # self.fill_mkb_items('d.xlsx')
        Scheme.objects.all().delete()
        self.fill_schemes('d.xlsx', "Дневной")
        self.fill_schemes('s.xlsx', "Круглосуточный")






        pass







if __name__ == '__main__':
    get_items('d.xlsx');
